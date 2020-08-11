import os
import re
from typing import Generator

import requests
from bs4 import BeautifulSoup as bs
from fake_useragent import UserAgent
from openpyxl import load_workbook
from progress.bar import IncrementalBar


base_dir = os.path.dirname(__file__)
xlsx_file = base_dir + '\\for_parsing.xlsx'
wb = load_workbook(xlsx_file)
sheet = wb.get_sheet_by_name('only_product')


def read_xlsx() -> Generator:
    """Читаем xlsx файл"""
    row_count = sheet.max_row
    print(f'{row_count - 1} urls уйдут на обход')
    bar = IncrementalBar('Processing', max = row_count - 1)
    for row in range(2, row_count + 1):
        url = sheet['A' + str(row)].value
        bar.next()
        yield url, row


def get_user_agent() -> dict:
    """Генерим user agent"""
    ua = UserAgent()
    header = {'user-agent': ua.random}
    return header


def get_response(url: str, header: dict) -> requests.Response:
    """Получаем http(s) ответ от сервера"""
    r = requests.get(url, headers=header, stream=True)
    if r.ok:
        return r
    else:
        print(f'Ошибка {r.status_code}')


def get_all_page_data(response: requests.Response) -> dict:
    """Сбор всех необходимых данных"""
    html = response.text
    soup = bs(html, 'lxml')
    print(response.url)

    try:
        price_sourse = soup.find('div', class_='price').text.strip()
        # Валюта
        price_currency = price_sourse.split()[-1]
        current_price = soup.find('span', itemprop='price').text
        price = current_price + ' ' + price_currency
    except AttributeError:
        price = ''

    try:
        old_price = soup.find('div', class_='old-price').text.strip()
    except AttributeError:
        old_price = ''

    try:
        description = soup.find('div', id='tab-description') \
            .find('p').find_parent('div', class_='col-md-6').text.strip()
        description = re.sub(r' +', ' ', description)
        description = re.sub(r' \n', '\n', description)
        description = re.sub(r'\n ', '\n', description)
        description = re.sub(r'\n+', '\n', description)
    except AttributeError:
        description = ''

    try:
        ware_code = soup.find('span', class_='sku cod').text.strip()
        code = re.findall(r'\d+', ware_code)
    except AttributeError:
        code = ['']

    try:
        image_url = soup.find('meta', property='og:image')['content']
    except AttributeError:
        image_url = ''

    if image_url:
        img_name = image_url.split('/')[-1]
        img_path = '/images/' + img_name
    else:
        img_path = ''

    try:
        category_paths = soup.find_all('ol', class_='breadcrumb')
        category_path = ' > '.join([i.text.strip() for i in category_paths])
        category_path = re.sub(r'\n+', r'\n', category_path)
        category_path = re.sub(r'\n', ' > ', category_path)
    except AttributeError:
        category_path = ''

    try:
        characteristics = {}
        characteristics_trs = soup.find('table', class_='reviewtab').find_all('tr')
        if characteristics_trs:
            for chars in characteristics_trs:
                tds = chars.find_all('td')
                char_name = tds[0].text.strip()
                char_value = tds[1].text.strip()
                characteristics[char_name] = char_value
    except AttributeError:
        characteristics = {}

    page_data = {
        'price': price,
        'old_price': old_price,
        'description': description,
        'characteristics': characteristics,
        'category_path': category_path,
        'ware_code': code,
        'image_url': image_url,
        'img_path': img_path,
    }

    return page_data


def write_xlsx(data: dict, row: int) -> None:
    """Записываем данные в эксель файл"""
    sheet['I' + str(row)].value = data['price']
    sheet['J' + str(row)].value = data['old_price']
    sheet['K' + str(row)].value = data['description']
    sheet['L' + str(row)].value = data['category_path']
    if len(data['ware_code']) > 1:
        print('Два идентификатора одного товара?')
    sheet['M' + str(row)].value = data['ware_code'][0]
    sheet['N' + str(row)].value = data['image_url']
    sheet['O' + str(row)].value = data['img_path']
    characteristics = data['characteristics']
    if characteristics:
        col = 16
        for key, value in characteristics.items():
            characteristic = 'Название характеристики = ' + key + ' @@ Значение характеристики = ' + value
            sheet.cell(row=row, column=col).value = characteristic
            col += 1
    wb.save(xlsx_file)


def save_image(url: str):
    """Сохраняем картинку в локальное храниелище"""
    img_name = url.split('/')[-1]
    img_path = base_dir + '\\images\\' + img_name
    print('Качаем картинку')
    file_object = get_response(url, get_user_agent())
    with open(img_path, 'bw') as f:
        for chunk in file_object.iter_content(8192):
            f.write(chunk)


def main():
    """Основная функция модуля. Вызов всех необходимых функций"""
    user_agent = get_user_agent()
    xlsx_data = read_xlsx()
    for url, row in xlsx_data:
        response = get_response(url, user_agent)
        page_data = get_all_page_data(response)
        write_xlsx(page_data, row)
        image_url = page_data['image_url']
        if image_url:
            save_image(image_url)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(e)
